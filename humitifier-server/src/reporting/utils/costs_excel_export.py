from datetime import datetime, date, timezone
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from io import BytesIO

from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from hosts.models import Host
from humitifier_common.artefacts import Hardware
from reporting.models import CostsScheme
from reporting.utils import CostsBreakdown, calculate_from_hardware_artefact
from reporting.utils.get_server_hardware import (
    Server,
    get_hardware_fact,
    get_hardware_for_hosts,
)


##
## Data models
##


class _Host(BaseModel):
    hostname: str
    hardware: Hardware
    cost_breakdown: CostsBreakdown
    scan_date: datetime
    month: datetime


class CostExcelCustomer(BaseModel):
    name: str | None = None
    hosts: list[_Host]


##
## Data-based generator funcs
##


# Not used anymore; left here for a simpler version of the calculation
def create_current_cost_excel(
    costs_scheme: CostsScheme, filename: str, customers: list[str] | str | None = None
) -> BytesIO:
    customer_info = []

    if not customers:
        customers = _get_all_customers()

    for customer in customers:
        hosts = Host.objects.filter(customer=customer, archived=False)

        server_details = get_hardware_for_hosts(hosts)

        hosts = []
        for server in server_details:
            cost_breakdown = calculate_from_hardware_artefact(
                server.hardware,
                costs_scheme,
            )

            hosts.append(
                _Host(
                    hostname=server.hostname,
                    hardware=server.hardware,
                    cost_breakdown=cost_breakdown,
                    scan_date=server.scan_date,
                    month=server.scan_date,
                )
            )

        customer_info.append(
            CostExcelCustomer(
                name=customer,
                hosts=hosts,
            )
        )

    return generate_excel(filename, customer_info)


def create_timeseries_cost_excel(
    costs_scheme: CostsScheme,
    filename: str,
    start_date: date,
    end_date: date,
    customers: list[str] | str | None = None,
):
    customer_info = []

    months = _get_months_between_dates(start_date, end_date)

    if not customers:
        customers = _get_all_customers()

    for customer in customers:
        # Get all billable for customer
        hosts = Host.objects.filter(customer=customer, billable=True)
        # Filter out all servers archived before our start date
        hosts = hosts.exclude(archival_date__lt=start_date)
        # Filter out all servers created after our end date
        hosts = hosts.exclude(created_at__gt=end_date)

        entries = []
        for host in hosts:
            for month in months:
                info = _get_server_info_for_month(host, month)

                if not info:
                    continue

                cost_breakdown = calculate_from_hardware_artefact(
                    info.hardware,
                    costs_scheme,
                )

                entries.append(
                    _Host(
                        hostname=info.hostname,
                        hardware=info.hardware,
                        cost_breakdown=cost_breakdown,
                        scan_date=info.scan_date,
                        month=month,
                    )
                )

        customer_info.append(
            CostExcelCustomer(
                name=customer,
                hosts=entries,
            )
        )

    return generate_excel(filename, customer_info)


##
## Actual Excel creation func
##


def generate_excel(
    filename: str, customers: list[CostExcelCustomer] | CostExcelCustomer
) -> BytesIO:
    wb = Workbook()
    main_sheet = wb.active

    TOTAL_ROWS = "5000"  # Previously this was 999; which was not enough...

    if not isinstance(customers, (list, tuple)):
        customers = [customers]

    header = NamedStyle(name="header")
    header.font = Font(name="Calibri", bold=True, color="000000")
    header.fill = PatternFill("solid", fgColor="DDDDDD")
    wb.add_named_style(header)

    for customer in customers:
        ws = wb.create_sheet(_get_customer_sheet_name(customer.name))
        ws.append(
            [
                "Server",
                "Billing month",
                "Num. CPU",
                "Memory",
                "Storage",
                "VM Costs",
                "Storage Costs",
                "Support Costs",
                "Total Costs",
            ]
        )

        for column in [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
        ]:
            cell = ws[f"{column}1"]
            cell.style = header

        total_server_costs = Decimal("0")
        total_support_costs = Decimal("0")

        for i, host in enumerate(customer.hosts):

            total_server_costs += host.cost_breakdown.total_hardware_costs
            total_support_costs += host.cost_breakdown.management

            # Excel doesn't know what timezones are
            billing_month = host.month.strftime("%B %Y")

            ws.append(
                [
                    host.hostname,
                    billing_month,
                    host.cost_breakdown.num_cpu,
                    host.cost_breakdown.memory_size,
                    host.cost_breakdown.total_storage_usage,
                    host.cost_breakdown.vm_costs,
                    host.cost_breakdown.total_storage_costs,
                    host.cost_breakdown.management,
                    host.cost_breakdown.total_costs,
                ]
            )

            ws[f"F{i+2}"].number_format = "€ #,##0.00"
            ws[f"G{i+2}"].number_format = "€ #,##0.00"
            ws[f"H{i+2}"].number_format = "€ #,##0.00"
            ws[f"I{i+2}"].number_format = "€ #,##0.00"

        _set_sheet_width(ws)

        max_index = len(customer.hosts) + 1

        filters = ws.auto_filter
        filters.ref = f"A1:I{max_index}"
        for column in [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
        ]:
            ws.auto_filter.add_sort_condition(f"{column}2:{column}{max_index}")

    main_sheet.title = "Overview"
    main_sheet.column_dimensions["A"].width = 30
    main_sheet.column_dimensions["B"].width = 20
    main_sheet.column_dimensions["C"].width = 20
    main_sheet.column_dimensions["D"].width = 20

    main_sheet.append(["Customer", "Server costs", "Support costs", "Total costs"])

    for column in ["A", "B", "C", "D"]:
        cell = main_sheet[f"{column}1"]
        cell.style = header

    for i, customer in enumerate(customers):
        sheet_name = _get_customer_sheet_name(customer.name)
        total_vm_costs = (
            f"=SUM('{sheet_name}'!F2:F{TOTAL_ROWS}) + SUM('{sheet_name}'!G2:"
            f"G{TOTAL_ROWS})"
        )
        total_support_costs = f"=ROUND(SUM('{sheet_name}'!H2:H{TOTAL_ROWS}),2)"
        total_costs = f"=ROUND(SUM('{sheet_name}'!I2:I{TOTAL_ROWS}),2)"
        main_sheet.append(
            [
                customer.name,
                total_vm_costs,
                total_support_costs,
                total_costs,
            ]
        )

        main_sheet[f"B{i + 2}"].number_format = "€ #,##0.00"
        main_sheet[f"C{i + 2}"].number_format = "€ #,##0.00"
        main_sheet[f"D{i + 2}"].number_format = "€ #,##0.00"

    buffer = BytesIO()
    buffer.name = filename
    wb.save(buffer)

    return buffer


##
## Helpers
##


def _get_all_customers():
    return Host.objects.values_list("customer", flat=True).order_by().distinct()


def _get_months_between_dates(start: date, end: date) -> list[datetime]:
    dates = []
    current = datetime(
        year=start.year,
        month=start.month,
        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
        tzinfo=timezone.utc,
    )
    end_date = datetime(
        year=end.year,
        month=end.month,
        day=1,
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
        tzinfo=timezone.utc,
    )

    while current <= end_date:
        dates.append(current)
        current += relativedelta(months=1)

    return dates


def _get_server_info_for_month(host: Host, month: datetime) -> Server | None:
    # First, see if the server was turned off for this month
    if host.offline_periods.filter(
        start_date__lte=month,
        end_date__gte=month,
    ).exists():
        # If so, we don't have any data
        return None

    # Second, check if this server was created after this month
    if host.created_at > month:
        # We can't bill this then :)
        return None

    # Third, check if this server was archived before this month
    if host.archival_date and host.archival_date < month:
        return None

    # Get the latest scan for this month
    # We do this by ignoring all scans before this month and then get the oldest
    # scan. This is a bit... weird, but this works around some edge cases with missing
    # data.
    scans = host.scans.exclude(created_at__lt=month).order_by("created_at")
    scans = scans.filter(data__version__gte=2)
    if not scans.exists():
        return None

    scan_for_month = scans.first().get_scan_object()

    return get_hardware_fact(scan_for_month)


def _get_customer_sheet_name(customer_name: str | None) -> str:
    if customer_name is None:
        customer_name = "Unknown"
    postfix = "- Server List"
    max_length = 31 - len(postfix) - 1  # Account for space and "-"
    truncated_customer = (
        customer_name[:max_length] if len(customer_name) > max_length else customer_name
    )
    return f"{truncated_customer} {postfix}"


def _set_sheet_width(ws: Worksheet):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 5
